"""Pandas read_file goldens: calamine vs openpyxl/xlrd/odf fallbacks."""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

import smt_processor

tests_path = os.path.dirname(os.path.realpath(__file__))
_ASSETS = Path(tests_path) / "assets"


def test_chinese_bom_datetime_header(tmp_path: Path) -> None:
    path = tmp_path / "cn_bom.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["title", datetime(2024, 6, 1, 12, 0, 0), "extra"])
    ws.append(["a", "b", "c"])
    ws.append(["d", "e", "f"])
    ws.append(["Designator", "Footprint", "Comment"])
    ws.append(["R1", "0402", "10k"])
    wb.save(path)
    df = smt_processor.read_file(str(path))
    cols = [str(c).strip() for c in df.columns]
    assert "Designator" in cols
    assert "R1" in df.iloc[:, 0].astype(str).values


def test_unnamed_merged_column_kept(tmp_path: Path) -> None:
    path = tmp_path / "merged.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:B1")
    ws["A1"] = "MergedHdr"
    ws["C1"] = "Keep"
    ws["A2"] = "left"
    ws["B2"] = "HASDATA"
    ws["C2"] = "right"
    wb.save(path)
    df = smt_processor.read_file(str(path))
    flat = df.astype(str).values.ravel().tolist()
    assert any("HASDATA" in str(v) for v in flat)
    assert len(df.columns) >= 2


def test_chip_size_text_0402(tmp_path: Path) -> None:
    path = tmp_path / "chip.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Footprint"
    ws["A2"] = "0402"
    ws["A2"].number_format = "@"
    ws["B1"] = "Ref"
    ws["B2"] = "R1"
    wb.save(path)
    df = smt_processor.read_file(str(path))
    fp = str(df.iloc[0]["Footprint"])
    assert "0402" in fp or fp.strip() == "402"


def test_formula_cached_or_fallback(tmp_path: Path) -> None:
    path = tmp_path / "formula.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "N"
    ws["A2"] = 2
    ws["B1"] = "Expr"
    ws["B2"] = "=1+1"
    wb.save(path)
    df = smt_processor.read_file(str(path))
    assert not df.empty
    n_col = str(df.iloc[0, 0])
    assert n_col in {"2", "2.0"}


def test_xlsx_that_is_csv_fallback(tmp_path: Path) -> None:
    path = tmp_path / "fake.xlsx"
    path.write_text("Designator,Footprint\nR1,0402\n", encoding="utf-8")
    df = smt_processor.read_file(str(path))
    blob = " ".join(df.astype(str).values.ravel())
    assert "R1" in blob


def test_header_none_vs_file_headers(tmp_path: Path) -> None:
    path = tmp_path / "hdr.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Designator", "Footprint"])
    ws.append(["C1", "0603"])
    wb.save(path)
    cli = smt_processor.read_file(str(path), column_headers_from_file=True)
    gui = smt_processor.read_file(str(path), column_headers_from_file=False)
    assert "Designator" in [str(c) for c in cli.columns]
    assert list(gui.columns) == ["0", "1"]
    assert "Designator" in gui.astype(str).values.ravel().tolist()


def test_read_file_ods_asset() -> None:
    path = _ASSETS / "bom.ods"
    if not path.is_file():
        pytest.skip("bom.ods missing")
    df = smt_processor.read_file(str(path))
    assert not df.empty


def test_read_file_xls_asset() -> None:
    path = _ASSETS / "bom.xls"
    if not path.is_file():
        pytest.skip("bom.xls missing")
    df = smt_processor.read_file(str(path))
    assert not df.empty


def test_calamine_matches_legacy_on_bom_xlsx() -> None:
    path = str(_ASSETS / "bom.xlsx")
    via_cal = pd.read_excel(path, engine="calamine")
    via_opy = pd.read_excel(path, engine="openpyxl")
    a = via_cal.astype(str).fillna("")
    b = via_opy.astype(str).fillna("")
    assert list(a.columns) == list(b.columns)
    assert a.shape == b.shape
    assert a.values.tolist() == b.values.tolist()
