#
# 2025-11-28
#

import logger
import datetime

# https://linuxhint.com/read-excel-file-python/
# https://openpyxl.readthedocs.io/en/stable/tutorial.html
import openpyxl

from pathlib import Path

from text_grid import TextGrid

# -----------------------------------------------------------------------------


def _require_existing_file(path: str) -> Path:
    assert path is not None
    if not str(path).strip():
        logger.error("Empty file path")
        raise FileNotFoundError("Empty file path")
    p = Path(path)
    if not p.is_file():
        logger.error("File not found: %s", path)
        raise FileNotFoundError(f"File not found: {path}")
    return p
def __check_row_valid(row_cells: list[str]) -> bool:
    # ignore rows with empty cells 'A,B,C' or cell 'A' with a long horizontal line
    row_valid = (len(row_cells) > 3) and (row_cells[0] or row_cells[1] or row_cells[2])
    row_valid = row_valid and not row_cells[0].startswith("___")
    return row_valid


def read_xlsx_sheet(path: str) -> TextGrid:
    """
    Reads entire sheet 0
    """
    p = _require_existing_file(path)
    logger.info(f"Reading file '{path}'")
    try:
        workbook = openpyxl.load_workbook(str(p))
    except Exception as e:
        logger.error("Cannot read XLSX file %s: %s", path, e)
        raise
    sheet = workbook.active
    tg = TextGrid()

    # Iterate the loop to read the cell values
    for row in sheet.iter_rows(
        min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True
    ):
        row_cells = []
        for cell in row:
            if cell is None:
                cell = ""
            else:
                if isinstance(cell, float) or isinstance(cell, int):
                    if isinstance(cell, float) and int(cell) == float(cell):
                        # prevent the conversion of '100' to '100.0'
                        cell = int(cell)
                    cell = repr(cell)
                elif isinstance(cell, datetime.datetime):
                    cell = str(cell)
            # change multiline cells into single-line
            cell = cell.replace("\n", " ⏎ ")
            row_cells.append(cell.strip())

        if __check_row_valid(row_cells):
            tg.rows_raw().append(row_cells)

    tg.nrows = len(tg.rows_raw())
    tg.ncols = sheet.max_column
    tg.align_number_of_columns()
    return tg
